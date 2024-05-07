import os
import asyncio

import openpyxl
from aiogram.fsm.context import FSMContext
import pandas as pd

from utils.database import db_settings_get_camb_w_num, db_settings_update_camb_w_num
from utils.states import Dictionary


def save_to_cambridge_dict(path, data: list) -> None:

    wb = openpyxl.load_workbook(path)
    sheet = wb['english']
    # idiom or not?
    # if data[0][0][1] == 'idiom':
    if len(data[0]) == 1:
        data = data[0]
        print('это идиома')
        if len(data[0][4]) != 0:
            sheet.append([data[0][0], data[0][1], data[0][2],
                        data[0][3], data[0][4][0]])
        else:
            sheet.append([data[0][0], data[0][1], data[0][2],
                        data[0][3]])
        wb.save(path)
    else:
        print('это слово')
        for mass in data[0]:
            print(mass)
            if len(mass) > 5:
                if len(mass[5]) > 0:
                    sheet.append([mass[0], mass[1],
                                mass[3], mass[4], mass[5][0]])
            else:
                try:
                    sheet.append([mass[0], mass[1],
                                mass[3], mass[4]])
                except IndexError:
                    if len(data[0][4]) != 0:
                        sheet.append([mass[0][0], mass[0][1], mass[0][2],
                                    mass[0][3], mass[0][4][0]])
                    else:
                        sheet.append([mass[0][0], mass[0][1], mass[0][2],
                                    mass[0][3]])
            wb.save(path)
    wb.save(path)
    
    
async def delete_from_cambridge_dict(path, data: list, user_id) -> None:
    wb = openpyxl.load_workbook(path)
    sheet = wb['english']
    col = sheet['A']
    val_col = []
    for i in col:
        val_col.append(i.value)

    print(data)
    if data == []:
        pnum = await db_settings_get_camb_w_num(user_id)
        if int(pnum[0][0])+1 >= len(col):
            pnum = 2
        else: 
            pnum = int(pnum[0][0])
        sheet.delete_rows(pnum)
        print(f'{pnum} row has deleted')
    else:
        for mass in data[0]:
            wb = openpyxl.load_workbook(path)
            sheet = wb['english']
            col = sheet['A']
            val_col = []
            for i in col:
                val_col.append(i.value)
            print(val_col.index(mass[0]))
            sheet.delete_rows(val_col.index(mass[0])+1)
            wb.save(path)
    wb.save(path)


def cambridge_dict_first_set(path, state: FSMContext) -> None:

    # df = pd.read_excel("./data/cambridge_dict.xlsx")
    # dfr = pd.DataFrame(df)
    # # dfr.to_excel(path, sheet_name='english')
    # file = open(path, "x")
    # dfr.to_excel(
    #     file,
    #     sheet_name='english'
    # )
    # file.close()

    df = pd.DataFrame([["", "", "", "", ""]], index=None, columns=[
                      "word", "part_of_speech", "translation", "meaning", "example"])
    with pd.ExcelWriter(path) as writer:
        df.to_excel(writer, index=None, sheet_name='english')

    print('УРААА🔥')
    wb = openpyxl.load_workbook(path)
    sheet = wb['english']
    sheet.delete_rows(2)
    wb.save(path)
    # cell = sheet['A1']
    # cell.value = 'word'
    # cell = sheet['B1']
    # cell.value = 'part_of_speech'
    # cell = sheet['C1']
    # cell.value = 'translation'
    # cell = sheet['D1']
    # cell.value = 'meaning'
    # cell = sheet['E1']
    # cell.value = 'example'

    # await state.set_state(Dictionary.cambridge_word_number)
    # await state.update_data(cambridge_word_number=1)


def does_word_exists_in_dict(word, path):
    wb = openpyxl.load_workbook(path)
    sheet = wb['english']
    col = sheet['A']

    if word in [i.value for i in col]:
        return True
    else:
        return False


async def cambridge_dict_iterator(path, user_id):
    pnum = await db_settings_get_camb_w_num(user_id)
    print(pnum)

    wb = openpyxl.load_workbook(path)
    sheet = wb['english']
    col = sheet['A']

    if int(pnum[0][0])+1 > len(col):
        pnum = 2
    else: 
        pnum = int(pnum[0][0])+1

    cell_range = sheet[f'A{pnum}':f'E{pnum}']

    # await state.update_data(cambridge_word_number=pnum+1)
    await db_settings_update_camb_w_num(user_id, int(pnum))

    return_data = []
    for i in cell_range[0]:
        return_data.append(i.value)

    return return_data
